import openpyxl

class AntropometriSK:
    def __init__(self, filepath):
        self.wb = openpyxl.load_workbook(filepath, data_only=True)

    def _to_float(self, value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            text = str(value).replace(",", ".").strip()
            return float(text)
        except ValueError:
            return None

    def _find_best_row_by_first_col(self, ws, target_value, start_row=1):
        best_row = None
        best_diff = float("inf")

        for r in range(start_row, ws.max_row + 1):
            cell_val = self._to_float(ws.cell(r, 1).value)
            if cell_val is None:
                continue

            diff = abs(cell_val - target_value)
            if diff < best_diff:
                best_diff = diff
                best_row = r

        return best_row

    def classify_tbu(self, jenis_kelamin, umur_bulan, tinggi_cm):
        sheet_name = "TBU L" if jenis_kelamin == "L" else "TBU P"
        ws = self.wb[sheet_name]

        row = self._find_best_row_by_first_col(ws, umur_bulan, start_row=5)
        if row is None:
            return "Data TBU tidak ditemukan"

        minus_3sd = self._to_float(ws.cell(row, 2).value)
        minus_2sd = self._to_float(ws.cell(row, 3).value)
        plus_2sd = self._to_float(ws.cell(row, 7).value)

        if None in (minus_3sd, minus_2sd, plus_2sd):
            return "Data TBU tidak lengkap"

        if tinggi_cm < minus_3sd:
            return "Sangat Pendek"
        elif tinggi_cm < minus_2sd:
            return "Pendek"
        elif tinggi_cm <= plus_2sd:
            return "Normal"
        else:
            return "Tinggi"
        
    def classify_bbtb(self, jenis_kelamin, umur_bulan, tinggi_cm, berat_kg):
        if umur_bulan <= 24:
            sheet_name = "BBPB L" if jenis_kelamin == "L" else "BBPB P"
        else:
            sheet_name = "BBTB L" if jenis_kelamin == "L" else "BBTB P"

        ws = self.wb[sheet_name]

        row = self._find_best_row_by_first_col(
            ws,
            tinggi_cm,
            start_row=5
        )

        if row is None:
            return "Data BB/PB atau BB/TB tidak ditemukan"

        minus_3sd = self._to_float(ws.cell(row, 2).value)
        minus_2sd = self._to_float(ws.cell(row, 3).value)
        plus_2sd = self._to_float(ws.cell(row, 7).value)

        if None in (minus_3sd, minus_2sd, plus_2sd):
            return "Data BB/PB atau BB/TB tidak lengkap"

        if berat_kg < minus_3sd:
            return "Sangat Kurus"
        elif berat_kg < minus_2sd:
            return "Kurus"
        elif berat_kg <= plus_2sd:
            return "Normal"
        else:
            return "Gemuk"
